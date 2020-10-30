# coding: utf-8
import unicodedata

from openpyxl import Workbook
from openpyxl import load_workbook



print("====")
print("This script creates the javascript for the autocomplete.")
print("====")
print("")

elfile = input("elements filename? must be .xlsx and same format as for importing! ")

print ("...working!")

def make_autocomplete(elfile):


    wb = load_workbook(filename = elfile)



    #wb_obj = openpyxl.load_workbook(path) 
      
    # Get workbook active sheet object 
    # from the active attribute 
    sheet_obj = wb.active 
      
    # Cell objects also have row, column,  
    # and coordinate attributes that provide 
    # location information for the cell. 
      
    # Note: The first row or  
    # column integer is 1, not 0. 
      
    # Cell object is created by using  
    # sheet object's cell() method. 
    cell_obj = sheet_obj.cell(row = 2, column = 5) 
      
    # Print value of cell object  
    # using the value attribute 
    #print(cell_obj.value) 

    num = sheet_obj.max_row

    pl = []

    for x in range(1,num+1):
        term = sheet_obj.cell(row = x, column = 1) 
        element = sheet_obj.cell(row = x, column = 2) 
        synonyms = sheet_obj.cell(row = x, column = 3) 

        uri = "https://schema.space/metasat/"+element.value

        flat_j = '{value:"'+element.value+'",label:"'+term.value+'",uri:"'+uri+'"}'
        pl.append(flat_j)

        if (element.value).lower() != (term.value).lower():
            flat_k = '{value:"'+element.value+'",label:"'+element.value+'",uri:"'+uri+'"}'
            pl.append(flat_k)

        try:
            synonym = (synonyms.value).split(",")

            for syn in synonym:
                flat_j = '{value:"'+element.value+'",label:"'+syn.strip()+' ('+term.value+')'+'",uri:"'+uri+'"}'
                pl.append(flat_j)

            #print (synonym)
        except:
            #print("none")
            pass



                
    #joins this list of pairs into a string
    q = ','.join(pl).encode('utf-8').strip()
    #print (q)

    js_file = open("metasat_autocomplete.js", "w")

    #opening javascript code, from Alex
    js_file.write(str("(function(b){function a(d){return d.split(/,\s*/)}function c(d){return a(d).pop()}uat_json=["))

    #writes the string of pairs to the file

    m = q.decode(encoding='UTF-8')
    js_file.write(str(q))

    #closing javascript code, from Alex
    js_file.write('];b.widget("custom.uatAutocomplete",b.ui.autocomplete,{options:{source:uat_json,multi:false,minLength:3}')
    js_file.write(',_create:function(){if(this.options.multi===false){this._super()}else{this._super();this.element.bind')
    js_file.write('("keydown",function(d){if(d.keyCode===b.ui.keyCode.TAB&&b(this).data("ui-autocomplete").menu.active)')
    js_file.write('{d.preventDefault()}}).autocomplete({source:function(e,d){d(b.ui.autocomplete.filter(uat_json,c(e.term)))}')
    js_file.write(',focus:function(){return false},select:function(e,f){var d=a(this.value);d.pop();d.push(f.item.value);d.push')
    js_file.write('("");this.value=d.join(", ");return false}})}}})}(jQuery));')

    js_file.close()


make_autocomplete(elfile)
print ("finished!")

print("copy 'metasat_autocomplete.js' to main/website/static/website/js/")
