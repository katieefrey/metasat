# standard lib packages
import sys
import os
import requests
import json
import urllib.parse
import time
import csv
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE','main.settings')

import django
django.setup()

from django.conf import settings

from crosswalk.models import ExternalSchema, ExternalElement
from metasat.models import Element, Segment, ElementFamily

print("====")
print("You are working with the "+str(settings.DBLOC)+" database!")
print("====")
print("")

runwhat = input("What do you want to import?\n\n  1) Elements Only\n  2) Crosswalks Only\n  3) Elements AND Crosswalks\n\n  4) Remind me about the file format requirements\n\n     Any other key to exit\n\nPlease enter 1, 2, 3, or 4: ")

## add item functions

def add_Segment(data1):
    d, created = Segment.objects.get_or_create(segment=data1)
    return d

def add_Family(data1):
    d, created = ElementFamily.objects.get_or_create(family=data1)
    return d

def add_ExternalSchema(data1, data2, data3, data4):
    d, created = ExternalSchema.objects.get_or_create(identifier=data1, name=data2, lang=data3, desc=data4)
    return d

def add_ExternalElement(data1, data3, data4, data5):
    d, created = ExternalElement.objects.get_or_create(identifier=data1, url=data3, source_id=data4, metasatelement_id=data5)

    return d

def add_Element(data1, data2, data3, data4, data5, data6, segments, families):

    e1, created = Element.objects.get_or_create(identifier=data1, term=data2, desc=data3, synonym=data4, example=data5, source=data6)
    e1.save()

    for x in segments:
        e1.segment.add(x)

    for x in families:
        e1.family.add(x)
    e1.save()


# def update_Element(data1, data2, data3, data4, data5, data6, segments, families):
#     d = Element.objects.get(identifier=data1)
#     d.term = data2 
#     d.desc = data3
#     d.synonym = data4
#     d.example = data5
#     d.source = data6

#     d.segment.clear()
#     for x in segments:
#         d.segment.add(x)

#     d.family.clear()
#     for x in families:
#         d.family.add(x)
#     d.save()



######### adding elements

def importelements(elfile):

    Element.objects.all().delete()
    ElementFamily.objects.all().delete()
    Segment.objects.all().delete()

    from openpyxl import load_workbook

    wb = load_workbook(filename = elfile)
    sheet_obj = wb.active 
    cell_obj = sheet_obj.cell(row = 2, column = 5) 

    num = sheet_obj.max_row

    print (num)

    for x in range(1,num+1):

        term = sheet_obj.cell(row = x, column = 2)
        identifier = sheet_obj.cell(row = x, column = 1)
        families = sheet_obj.cell(row = x, column = 4) 
        segments = sheet_obj.cell(row = x, column = 5) 
        desc = sheet_obj.cell(row = x, column = 6) 
        example = sheet_obj.cell(row = x, column = 7)
        synonyms = sheet_obj.cell(row = x, column = 3)
        source = sheet_obj.cell(row = x, column = 8)

        print(term)

        familys = (families.value).split(", ") 
        fam = []
        for x in familys:
           z = add_Family(x.strip())
           fam.append(z)

        segs1 = (segments.value).split(", ")
        seg = []
        for x in segs1:
            z = add_Segment(x)
            seg.append(z)

        print (identifier.value)
        add_Element((identifier.value).strip(),term.value,desc.value,synonyms.value,example.value,source.value, seg, fam)



############ adding crosswalks
def importcrosswalks(cwfile):

    ExternalSchema.objects.all().delete()
    ExternalElement.objects.all().delete()

    import pandas

    with open(cwfile) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        header = 0
        id_list = []
        name_list = []
        lang_list = []
        desc_list = []

        for row in csv_reader:
            if header == 0:
                # entries in the first row should be the url safe identifiers
                for title in row:
                    id_list.append(title)
                header = 1

            elif header == 1:
                # entries in the second row should be schema names with diacritics
                for name in row:
                    name_list.append(name)
                header = 2

            elif header == 2:
                # entries in the third row should be schema languages
                for lang in row:
                    lang_list.append(lang)
                header = 3

            elif header == 3:
                #entries in the fourth row should be crosswalk descriptions
                for desc in row:
                    desc_list.append(desc)
                header = 4

            else:
                pass


        elementid = id_list[0]

        # skip first item of each list
        id_use = id_list[1:]
        name_use = name_list[1:]
        lang_use = lang_list[1:]
        desc_use = desc_list[1:]

        df = pandas.read_csv(cwfile, encoding = "ISO-8859-1", usecols=id_list, skiprows=[1,2,3,4])

        num = 0
        for y in id_use:
            add_ExternalSchema(y.replace(' ','-'), name_use[num], lang_use[num], desc_use[num])
            schema = ExternalSchema.objects.get(identifier=y.replace(' ','-'))
            print (y)
            count = 0
            for x in df[y]:
                if pandas.isnull(x):
                    pass
                else:
                    print(x)
                    print(df[elementid][count]) # elementid

                    if str(x).startswith("http"):
                        uri = x
                    else:
                        uri = None

                    elid = Element.objects.get(identifier=df[elementid][count])
                    add_ExternalElement(str(x), uri, schema.id, elid.id)
                count+=1
            num+=1


if runwhat == str(1):
    elfile = input("elements filename? must be .xlsx! ")
    importelements(elfile)
    print ("finished!")

elif runwhat == str(2):
    cwfile = input("crosswalk filename? must be .csv! ")
    importcrosswalks(cwfile)
    print ("finished!")

elif runwhat == str(3):
    elfile = input("elements filename? must be .xlsx! ")
    cwfile = input("crosswalk filename? must be .csv! ")
    importelements(elfile)
    importcrosswalks(cwfile)
    print ("finished!")

elif runwhat == str(4):
    print("")
    print("")
    print("===Elements File===")
    print("")
    print("Must be a Microsoft Excel .xlsx file.")
    print("Must have NOT have a header row")
    print("Must have columns in this order:")
    print("")
    print("    A) Identifier")
    print("        - must be URL safe: no spaces, no accent marks, no diacritics, no punctuation")
    print("    B) Term")
    print("    C) Synonyms")
    print("        - must be comma separated")
    print("    D) Families")
    print("        - must be comma separated")
    print("    E) Segments")
    print("    F) Description")
    print("    G) Examples")
    print("    H) Conditions (not used)")
    print("    I) Source")
    print("        - source of defintion")
    print("    J) Deprecated")
    print("        - 'yes' or blank")
    print("    K) Wiki source (not used")
    print("")
    print("")

    print("===Crosswalks File===")
    print("")
    print("Must be a comma separated .csv file.")
    print("Must have THREE have a header rows")
    print("the header rows must be in this order:")
    print("")
    print("    1) URL safe schema name")
    print("        - no accent marks, no diacritics, no punctuation")
    print("    2) Proper schema name")
    print("        - accent marks, diacritics, punctuation are fine")
    print("    3) Schema name language")
    print("        - the language the schema name is written in, to improve assistive technologies")
    print("")
    print("After these three header rows, the first column must contain the identifier of the Element being crosswalked to.")
    print("All other colunmns should have the URI or other identifier of the terms in the crosswalked scheme or be left BLANK if no match exists.")
    print("Include http in all eternal element URIs, if no http included, the identifier will not be treated as a URI.")
    print("")
    print("")
    print("See the Example Files in the metasat github directory if needed")
    print("")
    print("")
   

else:
    print("invalid entry rerun script")


